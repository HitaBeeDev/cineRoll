import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 8th Academy Awards (1936) — films released in 1935
# Columns: Id | Award Year | Movie Name | Release Year | Film Type | Type Of Award | Award Winner | Award Nominee

data = [

    # ── Actor ─────────────────────────────────────────────────────────────────
    ('OSC-1936-40', 1936, 'The Informer',        1935, 'movie', 'Actor', 'Victor McLaglen', 'Victor McLaglen'),
    ('OSC-1936-09', 1936, 'Black Fury',           1935, 'movie', 'Actor', 'NaN', 'Paul Muni'),
    ('OSC-1936-27', 1936, 'Mutiny on the Bounty', 1935, 'movie', 'Actor', 'NaN', 'Clark Gable'),
    ('OSC-1936-27', 1936, 'Mutiny on the Bounty', 1935, 'movie', 'Actor', 'NaN', 'Charles Laughton'),
    ('OSC-1936-27', 1936, 'Mutiny on the Bounty', 1935, 'movie', 'Actor', 'NaN', 'Franchot Tone'),

    # ── Actress ───────────────────────────────────────────────────────────────
    ('OSC-1936-15', 1936, 'Dangerous',       1935, 'movie', 'Actress', 'Bette Davis', 'Bette Davis'),
    ('OSC-1936-03', 1936, 'Alice Adams',     1935, 'movie', 'Actress', 'NaN', 'Katharine Hepburn'),
    ('OSC-1936-07', 1936, 'Becky Sharp',     1935, 'movie', 'Actress', 'NaN', 'Miriam Hopkins'),
    ('OSC-1936-17', 1936, 'Escape Me Never', 1935, 'movie', 'Actress', 'NaN', 'Elisabeth Bergner'),
    ('OSC-1936-38', 1936, 'The Dark Angel',  1935, 'movie', 'Actress', 'NaN', 'Merle Oberon'),
    ('OSC-1936-31', 1936, 'Private Worlds',  1935, 'movie', 'Actress', 'NaN', 'Claudette Colbert'),

    # ── Art Direction ─────────────────────────────────────────────────────────
    ('OSC-1936-38', 1936, 'The Dark Angel',               1935, 'movie', 'Art Direction', 'Richard Day', 'Richard Day'),
    ('OSC-1936-41', 1936, 'The Lives of a Bengal Lancer', 1935, 'movie', 'Art Direction', 'NaN', 'Hans Dreier, Roland Anderson'),
    ('OSC-1936-45', 1936, 'Top Hat',                      1935, 'movie', 'Art Direction', 'NaN', 'Van Nest Polglase, Carroll Clark'),

    # ── Assistant Director ────────────────────────────────────────────────────
    ('OSC-1936-41', 1936, 'The Lives of a Bengal Lancer', 1935, 'movie', 'Assistant Director', 'Clem Beauchamp, Paul Wing', 'Clem Beauchamp, Paul Wing'),
    ('OSC-1936-16', 1936, 'David Copperfield',            1935, 'movie', 'Assistant Director', 'NaN', 'Joseph Newman'),
    ('OSC-1936-25', 1936, 'Les Misérables',               1935, 'movie', 'Assistant Director', 'NaN', 'Eric Stacey'),
    ('OSC-1936-02', 1936, 'A Midsummer Night\'s Dream',   1935, 'movie', 'Assistant Director', 'NaN', 'Sherry Shourds'),

    # ── Cinematography ────────────────────────────────────────────────────────
    ('OSC-1936-02', 1936, 'A Midsummer Night\'s Dream', 1935, 'movie', 'Cinematography', 'Hal Mohr', 'Hal Mohr'),
    ('OSC-1936-06', 1936, 'Barbary Coast',               1935, 'movie', 'Cinematography', 'NaN', 'Ray June'),
    ('OSC-1936-37', 1936, 'The Crusades',                1935, 'movie', 'Cinematography', 'NaN', 'Victor Milner'),
    ('OSC-1936-25', 1936, 'Les Misérables',              1935, 'movie', 'Cinematography', 'NaN', 'Gregg Toland'),

    # ── Dance Direction ───────────────────────────────────────────────────────
    ('OSC-1936-12', 1936, 'Broadway Melody of 1936', 1935, 'movie', 'Dance Direction',
     '"I\'ve Got a Feeling You\'re Fooling" from "Broadway Melody of 1936"',
     '"I\'ve Got a Feeling You\'re Fooling" from "Broadway Melody of 1936"'),
    ('OSC-1936-18', 1936, 'Folies Bergere', 1935, 'movie', 'Dance Direction',
     '"Straw Hat" from "Folies Bergere"',
     '"Straw Hat" from "Folies Bergere"'),
    ('OSC-1936-04', 1936, 'All the King\'s Horses', 1935, 'movie', 'Dance Direction',
     'NaN', '"Viennese Waltz" from "All the King\'s Horses"'),
    ('OSC-1936-08', 1936, 'Big Broadcast of 1936', 1935, 'movie', 'Dance Direction',
     'NaN', '"It\'s the Animal in Me" from "Big Broadcast of 1936"'),
    ('OSC-1936-11', 1936, 'Broadway Hostess', 1935, 'movie', 'Dance Direction',
     'NaN', '"Playboy from Paree" from "Broadway Hostess"'),
    ('OSC-1936-20', 1936, 'Go into Your Dance', 1935, 'movie', 'Dance Direction',
     'NaN', '"Latin from Manhattan" from "Go into Your Dance"'),
    ('OSC-1936-21', 1936, 'Gold Diggers of 1935', 1935, 'movie', 'Dance Direction',
     'NaN', '"Lullaby of Broadway" from "Gold Diggers of 1935", "The Words Are In My Heart" from "Gold Diggers of 1935"'),
    ('OSC-1936-24', 1936, 'King of Burlesque', 1935, 'movie', 'Dance Direction',
     'NaN', '"Lovely Lady" from "King of Burlesque", "Too Good To Be True" from "King of Burlesque"'),
    ('OSC-1936-34', 1936, 'She', 1935, 'movie', 'Dance Direction',
     'NaN', '"Hall of Kings" from "She"'),
    ('OSC-1936-45', 1936, 'Top Hat', 1935, 'movie', 'Dance Direction',
     'NaN', '"Piccolino" from "Top Hat", "Top Hat, White Tie, and Tails" from "Top Hat"'),

    # ── Directing ─────────────────────────────────────────────────────────────
    ('OSC-1936-40', 1936, 'The Informer',               1935, 'movie', 'Directing', 'John Ford', 'John Ford'),
    ('OSC-1936-14', 1936, 'Captain Blood',               1935, 'movie', 'Directing', 'NaN', 'Michael Curtiz'),
    ('OSC-1936-27', 1936, 'Mutiny on the Bounty',        1935, 'movie', 'Directing', 'NaN', 'Frank Lloyd'),
    ('OSC-1936-41', 1936, 'The Lives of a Bengal Lancer',1935, 'movie', 'Directing', 'NaN', 'Henry Hathaway'),

    # ── Film Editing ──────────────────────────────────────────────────────────
    ('OSC-1936-02', 1936, 'A Midsummer Night\'s Dream',  1935, 'movie', 'Film Editing', 'Ralph Dawson', 'Ralph Dawson'),
    ('OSC-1936-16', 1936, 'David Copperfield',            1935, 'movie', 'Film Editing', 'NaN', 'Robert J. Kern'),
    ('OSC-1936-25', 1936, 'Les Misérables',               1935, 'movie', 'Film Editing', 'NaN', 'Barbara McLean'),
    ('OSC-1936-27', 1936, 'Mutiny on the Bounty',         1935, 'movie', 'Film Editing', 'NaN', 'Margaret Booth'),
    ('OSC-1936-40', 1936, 'The Informer',                 1935, 'movie', 'Film Editing', 'NaN', 'George Hively'),
    ('OSC-1936-41', 1936, 'The Lives of a Bengal Lancer', 1935, 'movie', 'Film Editing', 'NaN', 'Ellsworth Hoagland'),

    # ── Music (Scoring) ───────────────────────────────────────────────────────
    ('OSC-1936-40', 1936, 'The Informer', 1935, 'movie', 'Music (Scoring)',
     'RKO Radio Studio Music Department, Max Steiner, head of department (Score by Max Steiner)',
     'RKO Radio Studio Music Department, Max Steiner, head of department (Score by Max Steiner)'),
    ('OSC-1936-14', 1936, 'Captain Blood', 1935, 'movie', 'Music (Scoring)',
     'NaN', 'Warner Bros.-First National Studio Music Department, Leo Forbstein, head of department (Score by Erich Wolfgang Korngold)'),
    ('OSC-1936-27', 1936, 'Mutiny on the Bounty', 1935, 'movie', 'Music (Scoring)',
     'NaN', 'Metro-Goldwyn-Mayer Studio Music Department, Nat W. Finston, head of department (Score by Herbert Stothart)'),
    ('OSC-1936-30', 1936, 'Peter Ibbetson', 1935, 'movie', 'Music (Scoring)',
     'NaN', 'Paramount Studio Music Department, Irvin Talbot, head of department (Score by Ernst Toch)'),

    # ── Music (Song) ──────────────────────────────────────────────────────────
    ('OSC-1936-21', 1936, 'Gold Diggers of 1935', 1935, 'movie', 'Music (Song)',
     'Lullaby Of Broadway in "Gold Diggers of 1935" Music by Harry Warren; Lyrics by Al Dubin',
     'Lullaby Of Broadway in "Gold Diggers of 1935" Music by Harry Warren; Lyrics by Al Dubin'),
    ('OSC-1936-32', 1936, 'Roberta', 1935, 'movie', 'Music (Song)',
     'NaN', 'Lovely To Look At in "Roberta" Music by Jerome Kern; Lyrics by Dorothy Fields and Jimmy McHugh'),
    ('OSC-1936-45', 1936, 'Top Hat', 1935, 'movie', 'Music (Song)',
     'NaN', 'Cheek To Cheek in "Top Hat" Music and Lyrics by Irving Berlin'),

    # ── Outstanding Production ────────────────────────────────────────────────
    ('OSC-1936-27', 1936, 'Mutiny on the Bounty',        1935, 'movie', 'Outstanding Production', 'Metro-Goldwyn-Mayer', 'Metro-Goldwyn-Mayer'),
    ('OSC-1936-03', 1936, 'Alice Adams',                  1935, 'movie', 'Outstanding Production', 'NaN', 'RKO Radio'),
    ('OSC-1936-12', 1936, 'Broadway Melody of 1936',      1935, 'movie', 'Outstanding Production', 'NaN', 'Metro-Goldwyn-Mayer'),
    ('OSC-1936-14', 1936, 'Captain Blood',                1935, 'movie', 'Outstanding Production', 'NaN', 'Cosmopolitan'),
    ('OSC-1936-16', 1936, 'David Copperfield',            1935, 'movie', 'Outstanding Production', 'NaN', 'Metro-Goldwyn-Mayer'),
    ('OSC-1936-25', 1936, 'Les Misérables',               1935, 'movie', 'Outstanding Production', 'NaN', '20th Century'),
    ('OSC-1936-41', 1936, 'The Lives of a Bengal Lancer', 1935, 'movie', 'Outstanding Production', 'NaN', 'Paramount'),
    ('OSC-1936-02', 1936, 'A Midsummer Night\'s Dream',   1935, 'movie', 'Outstanding Production', 'NaN', 'Warner Bros.'),
    ('OSC-1936-28', 1936, 'Naughty Marietta',             1935, 'movie', 'Outstanding Production', 'NaN', 'Metro-Goldwyn-Mayer'),
    ('OSC-1936-33', 1936, 'Ruggles of Red Gap',           1935, 'movie', 'Outstanding Production', 'NaN', 'Paramount'),
    ('OSC-1936-40', 1936, 'The Informer',                 1935, 'movie', 'Outstanding Production', 'NaN', 'RKO Radio'),
    ('OSC-1936-45', 1936, 'Top Hat',                      1935, 'movie', 'Outstanding Production', 'NaN', 'RKO Radio'),

    # ── Short Subject (Cartoon) ───────────────────────────────────────────────
    ('OSC-1936-43', 1936, 'Three Orphan Kittens',  1935, 'Animated Short Film', 'Short Subject (Cartoon)', 'Walt Disney, Producer', 'Walt Disney, Producer'),
    ('OSC-1936-36', 1936, 'The Calico Dragon',     1935, 'Animated Short Film', 'Short Subject (Cartoon)', 'NaN', 'Harman-Ising'),
    ('OSC-1936-46', 1936, 'Who Killed Cock Robin?',1935, 'Animated Short Film', 'Short Subject (Cartoon)', 'NaN', 'Walt Disney, Producer'),

    # ── Short Subject (Comedy) ────────────────────────────────────────────────
    ('OSC-1936-22', 1936, 'How to Sleep',  1935, 'Live Action Short Film', 'Short Subject (Comedy)', 'Jack Chertok, Producer', 'Jack Chertok, Producer'),
    ('OSC-1936-29', 1936, 'Oh, My Nerves', 1935, 'Live Action Short Film', 'Short Subject (Comedy)', 'NaN', 'Jules White, Producer'),
    ('OSC-1936-44', 1936, 'Tit for Tat',   1935, 'Live Action Short Film', 'Short Subject (Comedy)', 'NaN', 'Hal Roach, Producer'),

    # ── Short Subject (Novelty) ───────────────────────────────────────────────
    ('OSC-1936-47', 1936, 'Wings over Mt. Everest', 1935, 'Live Action Short Film', 'Short Subject (Novelty)', 'Gaumont British and Skibo Productions', 'Gaumont British and Skibo Productions'),
    ('OSC-1936-05', 1936, 'Audioscopiks',            1935, 'Live Action Short Film', 'Short Subject (Novelty)', 'NaN', 'Pete Smith, Producer'),
    ('OSC-1936-13', 1936, 'Camera Thrills',          1935, 'Live Action Short Film', 'Short Subject (Novelty)', 'NaN', 'Universal'),

    # ── Sound Recording ───────────────────────────────────────────────────────
    ('OSC-1936-28', 1936, 'Naughty Marietta',             1935, 'movie', 'Sound Recording', 'Metro-Goldwyn-Mayer Studio Sound Department, Douglas Shearer, Sound Director', 'Metro-Goldwyn-Mayer Studio Sound Department, Douglas Shearer, Sound Director'),
    ('OSC-1936-01', 1936, '$1,000 a Minute',              1935, 'movie', 'Sound Recording', 'NaN', 'Republic Studio Sound Department'),
    ('OSC-1936-10', 1936, 'Bride of Frankenstein',        1935, 'movie', 'Sound Recording', 'NaN', 'Universal Studio Sound Department, Gilbert Kurland, Sound Director'),
    ('OSC-1936-14', 1936, 'Captain Blood',                1935, 'movie', 'Sound Recording', 'NaN', 'Warner Bros.-First National Studio Sound Department, Nathan Levinson, Sound Director'),
    ('OSC-1936-38', 1936, 'The Dark Angel',               1935, 'movie', 'Sound Recording', 'NaN', 'United Artists Studio Sound Department, Thomas T. Moulton, Sound Director'),
    ('OSC-1936-23', 1936, 'I Dream Too Much',             1935, 'movie', 'Sound Recording', 'NaN', 'RKO Radio Studio Sound Department, Carl Dreher, Sound Director'),
    ('OSC-1936-41', 1936, 'The Lives of a Bengal Lancer', 1935, 'movie', 'Sound Recording', 'NaN', 'Paramount Studio Sound Department, Franklin B. Hansen, Sound Director'),
    ('OSC-1936-26', 1936, 'Love Me Forever',              1935, 'movie', 'Sound Recording', 'NaN', 'Columbia Studio Sound Department, John Livadary, Sound Director'),
    ('OSC-1936-35', 1936, 'Thanks a Million',             1935, 'movie', 'Sound Recording', 'NaN', '20th Century-Fox Studio Sound Department, E. H. Hansen, Sound Director'),

    # ── Writing (Original Story) ──────────────────────────────────────────────
    ('OSC-1936-42', 1936, 'The Scoundrel',          1935, 'movie', 'Writing (Original Story)', 'Ben Hecht, Charles MacArthur', 'Ben Hecht, Charles MacArthur'),
    ('OSC-1936-12', 1936, 'Broadway Melody of 1936',1935, 'movie', 'Writing (Original Story)', 'NaN', 'Moss Hart'),
    ('OSC-1936-19', 1936, 'G-Men',                  1935, 'movie', 'Writing (Original Story)', 'NaN', 'Gregory Rogers'),
    ('OSC-1936-39', 1936, 'The Gay Deception',      1935, 'movie', 'Writing (Original Story)', 'NaN', 'Don Hartman, Stephen Avery'),

    # ── Writing (Screenplay) ──────────────────────────────────────────────────
    ('OSC-1936-40', 1936, 'The Informer',               1935, 'movie', 'Writing (Screenplay)', 'Dudley Nichols', 'Dudley Nichols'),
    ('OSC-1936-14', 1936, 'Captain Blood',               1935, 'movie', 'Writing (Screenplay)', 'NaN', 'Casey Robinson'),
    ('OSC-1936-27', 1936, 'Mutiny on the Bounty',        1935, 'movie', 'Writing (Screenplay)', 'NaN', 'Talbot Jennings, Jules Furthman, Carey Wilson'),
    ('OSC-1936-41', 1936, 'The Lives of a Bengal Lancer',1935, 'movie', 'Writing (Screenplay)', 'NaN', 'Screenplay by Waldemar Young, John L. Balderston, Achmed Abdullah; Adaptation by Grover Jones, William Slavens McNutt'),
]

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'

headers = ['Id', 'Award Year', 'Movie Name', 'Release Year', 'Film Type',
           'Type Of Award', 'Award Winner', 'Award Nominee']

header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF')
center      = Alignment(horizontal='center', vertical='center')

ws.append(headers)
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center

for row in data:
    ws.append(list(row))

for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

path = 'backend/film-data/movie excel datas/oscar/oscars_1936_structured.xlsx'
wb.save(path)
print(f'Saved {len(data)} rows')

from collections import Counter
wb2 = openpyxl.load_workbook(path)
ws2 = wb2.active
awards = Counter(row[5] for row in ws2.iter_rows(min_row=2, values_only=True))
for k, v in sorted(awards.items()):
    print(f'  {k}: {v} rows')
